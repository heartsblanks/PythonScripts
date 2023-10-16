import openpyxl
from openpyxl.styles import PatternFill
import xml.etree.ElementTree as ET
from datetime import datetime

# Parse the JUnit XML report
tree = ET.parse('junit_report.xml')
root = tree.getroot()

# Load an existing Excel workbook or create a new one if it doesn't exist
try:
    workbook = openpyxl.load_workbook('test_report.xlsx')
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "JUnit Test Report"
    # Add headers for each column
    sheet.append(['Package', 'Class', 'Test Case', 'Status', 'Error/Failure Message', 'Previous Status', 'Updated Date', 'Time Taken'])

# Create fill patterns for different status colors
passed_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
failed_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
error_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Iterate through the XML and populate/update the Excel sheet
for testcase in root.iter('testcase'):
    package, classname = testcase.attrib['classname'].rsplit('.', 1)
    test_case_name = testcase.attrib['name']
    failure_element = testcase.find('failure')
    error_element = testcase.find('error')
    
    # Check if the test case already exists in the Excel sheet
    test_case_exists = False
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
        if [cell.value for cell in row] == [package, classname, test_case_name]:
            test_case_exists = True
            status_cell = row[3]
            error_message_cell = row[4]
            previous_status_cell = row[5]
            updated_date_cell = row[6]
            time_taken_cell = row[7]
            break

    # If the test case exists, update its status, error message, and date
    if test_case_exists:
        status_cell.value = "Failed" if failure_element is not None else ("Error" if error_element is not None else "Passed")
        error_message_cell.value = failure_element.attrib.get('message') if failure_element is not None else (error_element.attrib.get('message') if error_element is not None else '')
        previous_status_cell.value = previous_status_cell.value if previous_status_cell.value else status_cell.value
        updated_date_cell.value = datetime.now()
        time_taken_cell.value = testcase.attrib.get('time', '')

        # Apply conditional formatting for the "Status" column
        if status_cell.value == "Failed":
            status_cell.fill = failed_fill
        elif status_cell.value == "Error":
            status_cell.fill = error_fill
        else:
            status_cell.fill = passed_fill
    # If the test case is new, add it to the Excel sheet
    else:
        status = "Failed" if failure_element is not None else ("Error" if error_element is not None else "Passed")
        error_message = failure_element.attrib.get('message') if failure_element is not None else (error_element.attrib.get('message') if error_element is not None else '')
        updated_date = datetime.now()
        time_taken = testcase.attrib.get('time', '')
        # Add a row for the test case and its details
        row = [package, classname, test_case_name, status, error_message, status, updated_date, time_taken]
        sheet.append(row)

# Save the Excel file
workbook.save('test_report.xlsx')